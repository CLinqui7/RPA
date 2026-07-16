#!/usr/bin/env python3
from __future__ import annotations
import argparse, json, subprocess, tempfile, zipfile
from pathlib import Path
from openpyxl import load_workbook


def clean(v):
    return '' if v is None else str(v).strip()


def color_sig(c):
    if c is None: return None
    return (clean(c.type), clean(c.rgb), clean(c.indexed), clean(c.theme), clean(c.tint), clean(c.auto))


def side_sig(s):
    if s is None: return None
    return (clean(s.style), color_sig(s.color))


def style_sig(cell):
    f, fill, b, a, p = cell.font, cell.fill, cell.border, cell.alignment, cell.protection
    return (
        (clean(f.name), clean(f.sz), bool(f.b), bool(f.i), clean(f.u), bool(f.strike), color_sig(f.color), clean(f.vertAlign)),
        (clean(fill.fill_type), color_sig(fill.fgColor), color_sig(fill.bgColor)),
        (side_sig(b.left), side_sig(b.right), side_sig(b.top), side_sig(b.bottom), side_sig(b.diagonal), bool(b.diagonalUp), bool(b.diagonalDown)),
        (clean(a.horizontal), clean(a.vertical), clean(a.textRotation), bool(a.wrapText), bool(a.shrinkToFit), clean(a.indent)),
        clean(cell.number_format),
        (bool(p.locked), bool(p.hidden)),
    )


def anchor_sig(img):
    a = getattr(img, 'anchor', None)
    if isinstance(a, str):
        return ('cell', a, round(float(img.width or 0), 3), round(float(img.height or 0), 3))
    def m(x):
        if x is None: return None
        return (int(x.col), int(x.colOff), int(x.row), int(x.rowOff))
    return (type(a).__name__, m(getattr(a, '_from', None)), m(getattr(a, 'to', None)), round(float(img.width or 0), 3), round(float(img.height or 0), 3))


def sheet_sig(ws):
    styles, formulas = {}, {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.has_style: styles[cell.coordinate] = style_sig(cell)
            if isinstance(cell.value, str) and cell.value.startswith('='): formulas[cell.coordinate] = cell.value
    row_dims = {
        str(k):(v.height, bool(v.hidden), int(v.outlineLevel or 0), bool(v.collapsed))
        for k,v in ws.row_dimensions.items()
        if v.height is not None or v.hidden or v.outlineLevel or v.collapsed
    }
    col_dims = {
        str(k):(v.min, v.max, v.width, bool(v.hidden), int(v.outlineLevel or 0), bool(v.collapsed), bool(v.bestFit))
        for k,v in ws.column_dimensions.items()
        if v.width is not None or v.hidden or v.outlineLevel or v.collapsed or v.bestFit
    }
    ps = ws.page_setup
    pm = ws.page_margins
    return {
        'dimension': ws.calculate_dimension(),
        'merged': sorted(str(x) for x in ws.merged_cells.ranges),
        'freeze': clean(ws.freeze_panes),
        'print_area': clean(ws.print_area),
        'print_rows': clean(ws.print_title_rows),
        'print_cols': clean(ws.print_title_cols),
        'page_setup': (clean(ps.orientation), clean(ps.paperSize), clean(ps.scale), clean(ps.fitToWidth), clean(ps.fitToHeight), clean(ps.pageOrder)),
        'margins': (pm.left, pm.right, pm.top, pm.bottom, pm.header, pm.footer),
        'row_dims': row_dims,
        'col_dims': col_dims,
        'styles': styles,
        'formulas': formulas,
        'images': sorted(anchor_sig(x) for x in getattr(ws, '_images', [])),
        'grid': bool(ws.sheet_view.showGridLines),
        'zoom': ws.sheet_view.zoomScale,
        'state': clean(ws.sheet_state),
    }


def workbook_sig(path: Path):
    with zipfile.ZipFile(path) as z:
        bad = z.testzip()
        if bad: raise RuntimeError(f'CORRUPT_XLSX_MEMBER:{bad}')
    wb = load_workbook(path, data_only=False, keep_vba=path.suffix.lower()=='.xlsm')
    return {'sheet_names': list(wb.sheetnames), 'sheets': {ws.title: sheet_sig(ws) for ws in wb.worksheets}}


def find_template(catalog, suffix):
    expected = suffix.replace('\\','/').lower()
    for item in catalog.get('templates', []):
        best = item.get('best_sheet') or {}
        if item.get('error') or not best.get('table'): continue
        source = clean(item.get('source')).replace('\\','/').lower()
        if source.endswith(expected) or f'!{expected}' in source: return item
    return None


def payload(code):
    return {'header': {'customer_code':code,'order_no':'AUDIT-ONLY','order_date':'2026-07-13','start_date':'2026-07-14','cancel_date':'2026-07-20','store_code':'AUDIT','division_code':'MJ','terms_code':'X6','warehouse_code':'PE'}, 'lines':[{'style_code':'AUDITSTYLE','color_code':'AUD','customer_style':'AUDITSTYLE','customer_color':'AUDIT','customer_sku':'AUDITSKU','customer_upc':'000000000000','qty_total':1,'size_raw':'AUDIT','sales_price':None,'description':'FORMAT PRESERVATION AUDIT','qty_buckets':{'QTY_SZ1':1}}]}


def audit_one(code, profile, catalog, engine):
    out = {'customer_code':code,'source_suffix':profile.get('source_suffix'),'expected_dimension':profile.get('expected_dimension')}
    if not profile.get('source_suffix'):
        out['result']='INTENTIONALLY_BLOCKED_NO_APPROVED_TEMPLATE'; return out
    template = find_template(catalog, profile['source_suffix'])
    if not template:
        out['result']='APPROVED_TEMPLATE_NOT_FOUND'; return out
    src = Path(template['template_path']); out['template_path']=str(src)
    try: src_sig = workbook_sig(src)
    except Exception as e: out['result']='SOURCE_WORKBOOK_INVALID'; out['error']=str(e); return out
    with tempfile.TemporaryDirectory(prefix='a2000-v4772-') as td:
        td = Path(td); p = td/'payload.json'; g = td/f'{code}-AUDIT{src.suffix.lower()}'
        p.write_text(json.dumps(payload(code), indent=2), encoding='utf-8')
        r = subprocess.run(['python3',str(engine),'generate','--template',str(src),'--payload',str(p),'--output',str(g)], text=True, capture_output=True)
        if r.returncode != 0:
            out['result']='GENERATION_FAILED'; out['error']=(r.stderr or r.stdout)[-8000:]; return out
        try: gen_sig = workbook_sig(g)
        except Exception as e: out['result']='GENERATED_WORKBOOK_INVALID'; out['error']=str(e); return out
    diffs=[]
    if src_sig['sheet_names'] != gen_sig['sheet_names']: diffs.append('sheet_names')
    for name in src_sig['sheet_names']:
        if src_sig['sheets'].get(name) != gen_sig['sheets'].get(name):
            a,b = src_sig['sheets'].get(name,{}), gen_sig['sheets'].get(name,{})
            for key in sorted(set(a)|set(b)):
                if a.get(key) != b.get(key): diffs.append(f'{name}.{key}')
    primary = src_sig['sheet_names'][0] if src_sig['sheet_names'] else None
    actual_dim = src_sig['sheets'].get(primary,{}).get('dimension') if primary else None
    expected_ok = not profile.get('expected_dimension') or actual_dim == profile.get('expected_dimension')
    out.update({'source_dimension':actual_dim,'expected_dimension_valid':expected_ok,'differences':diffs,'difference_count':len(diffs)})
    out['result']='PASS' if not diffs and expected_ok else 'REVIEW_SEMANTIC_FORMAT'
    return out


def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--catalog',required=True); ap.add_argument('--registry',required=True); ap.add_argument('--engine',required=True); ap.add_argument('--output',required=True); args=ap.parse_args()
    catalog=json.loads(Path(args.catalog).read_text(encoding='utf-8')); registry=json.loads(Path(args.registry).read_text(encoding='utf-8'))
    results=[audit_one(code,p,catalog,Path(args.engine)) for code,p in registry['customers'].items()]
    approved=[x for x in results if x.get('source_suffix')]; passed=[x for x in approved if x['result']=='PASS']; review=[x for x in approved if x['result']!='PASS']
    report={'audit':'A2000_V4_7_7_2_SEMANTIC_CHECKLIST_FORMAT_AUDIT','customer_count':len(results),'approved_template_count':len(approved),'passed_template_count':len(passed),'review_template_count':len(review),'all_approved_templates_pass':len(approved)==len(passed),'results':results}
    Path(args.output).write_text(json.dumps(report,indent=2),encoding='utf-8')
    print('COPY THIS RESULT TO CHATGPT'); print('='*116)
    print('AUDIT=A2000_V4_7_7_2_SEMANTIC_CHECKLIST_FORMAT_AUDIT'); print(f'CUSTOMER_COUNT={len(results)}'); print(f'APPROVED_TEMPLATE_COUNT={len(approved)}'); print(f'PASSED_TEMPLATE_COUNT={len(passed)}'); print(f'REVIEW_TEMPLATE_COUNT={len(review)}'); print('ALL_APPROVED_TEMPLATES_PASS='+str(report['all_approved_templates_pass']).lower()); print('REVIEW_CUSTOMERS='+','.join(x['customer_code'] for x in review))
    for x in review: print(f"REVIEW_REASON={x['customer_code']}|RESULT={x['result']}|DIFFS={'|'.join(x.get('differences',[])[:12])}")
    print(f'REPORT={args.output}'); print('A2000_WRITES_PERFORMED=NO'); print('SUPABASE_WRITES_PERFORMED=NO'); print('='*116)
    raise SystemExit(0 if report['all_approved_templates_pass'] else 2)

if __name__=='__main__': main()
