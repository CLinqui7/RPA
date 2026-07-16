#!/usr/bin/env python3
import csv, re, sys
from pathlib import Path
try:
    from openpyxl import load_workbook
except Exception:
    print('ERROR: openpyxl is required. Run: python3 -m pip install openpyxl', file=sys.stderr)
    raise

def clean(v): return '' if v is None else str(v).replace('\u00a0',' ').strip()
def norm(v): return re.sub(r'[^A-Z0-9]', '', clean(v).upper())
def parse_csv_line(line): return next(csv.reader([line]))

CANCEL_OPEN_LINES_HEADER_ALIASES = {
    'CANCELOPENLINES',
    'CANCELOPENLINE',
    'CANCELOPENLNS',
    'CANCELOPENLN',
    'CANCELOPENORDERLINES',
    'CANCELOPENSALESORDERLINES',
    'CXLOPENLINES',
    'CXLOPENLINE',
    'CNCL​OPENLINES'.replace('\u200b', ''),
}

def customer_cancel_open_lines_value(row):
    exact = []
    semantic = []

    for key, value in (row or {}).items():
        key_clean = clean(key)
        token = norm(key_clean)
        if not token:
            continue

        item = (clean(value), key_clean)

        if token in CANCEL_OPEN_LINES_HEADER_ALIASES:
            exact.append(item)
            continue

        has_cancel = (
            'CANCEL' in token
            or 'CNCL' in token
            or 'CXL' in token
        )
        has_open = 'OPEN' in token
        has_line = (
            'LINE' in token
            or token.endswith('LN')
            or 'LNS' in token
        )

        if has_cancel and has_open and has_line:
            semantic.append(item)

    matches = exact or semantic

    if not matches:
        return '', ''

    nonempty = [
        item for item in matches
        if clean(item[0])
    ]

    selected = nonempty[0] if nonempty else matches[0]
    return clean(selected[0]), clean(selected[1])

def row_values_from_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        vals = [clean(v) for v in row]
        if not any(vals): continue
        if len(vals)==1 or (vals[0].count(',')>3 and not any(vals[1:])):
            try: yield parse_csv_line(vals[0])
            except Exception: yield [vals[0]]
        else:
            yield vals

def dict_rows_from_xlsx(path):
    it=iter(row_values_from_xlsx(path))
    try: header=[clean(x) for x in next(it)]
    except StopIteration: return
    for vals in it:
        if not any(clean(v) for v in vals): continue
        yield {h: clean(vals[i]) if i<len(vals) else '' for i,h in enumerate(header)}

def dict_rows_from_csv(path):
    raw=Path(path).read_bytes(); text=None
    for enc in ('utf-8-sig','latin1','cp1252'):
        try: text=raw.decode(enc); break
        except Exception: pass
    if text is None: text=raw.decode('latin1',errors='replace')
    rows=csv.reader(text.splitlines())
    try: header=[clean(x) for x in next(rows)]
    except StopIteration: return
    expected_columns=len(header)
    for vals in rows:
        if not any(clean(v) for v in vals): continue
        source_columns=len(vals)
        # Some official CSV exports contain unquoted commas inside text fields.
        # Never silently shift St Addr/City/State/Active into the wrong columns.
        # For malformed rows preserve only the structurally safe leading keys
        # (Customer, Store) and mark the rest as unavailable. Exact printed
        # STORE_NO can still be validated by key existence; address/active/WH
        # data from a shifted row is intentionally not trusted.
        if source_columns != expected_columns:
            row={h:'' for h in header}
            if expected_columns > 0 and source_columns > 0: row[header[0]]=clean(vals[0])
            if expected_columns > 1 and source_columns > 1: row[header[1]]=clean(vals[1])
            row['__source_row_status']='malformed_unquoted_csv_columns'
        else:
            row={h: clean(vals[i]) if i<source_columns else '' for i,h in enumerate(header)}
            row['__source_row_status']='ok'
        row['__source_column_count']=str(source_columns)
        row['__source_expected_column_count']=str(expected_columns)
        yield row

def iter_table(master_dir,*names):
    for name in names:
        p=Path(master_dir)/name
        if p.exists():
            print(f'Loading {p.name}...', flush=True)
            if p.suffix.lower()=='.csv': return dict_rows_from_csv(p), str(p)
            return dict_rows_from_xlsx(p), str(p)
    return iter(()), None

def write_csv(path, headers, rows):
    with Path(path).open('w', newline='', encoding='utf-8') as f:
        w=csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for row in rows:
            w.writerow({h: row.get(h,'') for h in headers})

def main():
    master_dir=Path(sys.argv[1] if len(sys.argv)>1 else 'api/masters').resolve()
    out_dir=master_dir/'cache'
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f'Building compact master cache from: {master_dir}', flush=True)

    customer_iter,_=iter_table(master_dir,'customer_master.csv.xlsx','customer_master.xlsx','customer_master.csv')
    store_iter,_=iter_table(master_dir,'stores_master.csv','stores_master.csv.xlsx','stores_master.xlsx')
    sku_iter,_=iter_table(master_dir,'VR_SKU.xlsx')
    sku_z_iter,_=iter_table(master_dir,'VR_SKU_Z.xlsx','VR_SKU_Z.csv.xlsx','VR_SKU_Z.csv')
    upc_iter,_=iter_table(master_dir,'VR_UPC_STYLE.xlsx')
    color_iter,_=iter_table(master_dir,'VR_COLOR.xlsx')
    whse_iter,_=iter_table(master_dir,'whse_master.csv','warehouse_master.csv','whse_master.xlsx','whse_master.csv.xlsx')

    print('Writing customers...', flush=True)
    customers=[]; c_count=0; cancel_open_lines_columns=set()
    for r in customer_iter:
        c_count += 1
        code=clean(r.get('Customer')).upper()
        if not code: continue
        cancel_open_lines, cancel_open_lines_source = customer_cancel_open_lines_value(r)
        if cancel_open_lines_source: cancel_open_lines_columns.add(cancel_open_lines_source)
        customers.append({'customer':code,'name':clean(r.get('Cust Name')),'name_norm':norm(r.get('Cust Name')),'terms':clean(r.get('Terms')),'terms_description':clean(r.get('Terms Description')),'ship_via':clean(r.get('Ship Via')),'def_wh':clean(r.get('Def Wh')),'div':clean(r.get('Div')),'active':clean(r.get('Active')),'addr1':clean(r.get('Addr 1')),'city':clean(r.get('City')),'state':clean(r.get('State')),'postal':clean(r.get('Postal')),'cancel_open_lines':cancel_open_lines,'cancel_open_lines_source':cancel_open_lines_source})
    write_csv(out_dir/'customers.csv', ['customer','name','name_norm','terms','terms_description','ship_via','def_wh','div','active','addr1','city','state','postal','cancel_open_lines','cancel_open_lines_source'], customers)

    print('Writing stores...', flush=True)
    stores=[]; s_count=0; malformed_store_rows=0
    for r in store_iter:
        s_count += 1
        customer=clean(r.get('Customer')).upper(); store=clean(r.get('Store')).upper()
        if not customer or not store: continue
        source_row_status=clean(r.get('__source_row_status')) or 'ok'
        if source_row_status != 'ok': malformed_store_rows += 1
        stores.append({'customer':customer,'store':store,'name':clean(r.get('St Name')),'addr1':clean(r.get('St Addr 1')),'city':clean(r.get('St City')),'state':clean(r.get('St State')),'postal':clean(r.get('St Postal')),'ship_via':clean(r.get('Ship Via')),'wh':clean(r.get('Wh')),'active':clean(r.get('Active')),'source_row_status':source_row_status})
    write_csv(out_dir/'stores.csv', ['customer','store','name','addr1','city','state','postal','ship_via','wh','active','source_row_status'], stores)

    print('Writing SKU compact...', flush=True)
    sku_headers=['style','clr','style_descr','clr_desc','clr_abbr','sku','sku_descr','scale','scale_abbr','div','customer','master_style','style_alias','invoice_descr','price','pack_qty','wh','style_norm','master_style_norm','style_alias_norm']
    sku_path=out_dir/'sku.csv'
    sku_count=0; known_styles=set()
    with sku_path.open('w', newline='', encoding='utf-8') as f:
        w=csv.DictWriter(f, fieldnames=sku_headers); w.writeheader()
        for r in sku_iter:
            sku_count += 1
            style=clean(r.get('Style')).upper(); clr=clean(r.get('Clr')).upper()
            if not style: continue
            known_styles.add(style)
            row={'style':style,'clr':clr,'style_descr':clean(r.get('Style Descr')),'clr_desc':clean(r.get('Clr Desc')),'clr_abbr':clean(r.get('Clr Abbr')),'sku':clean(r.get('Sku')),'sku_descr':clean(r.get('Sku Descr')),'scale':clean(r.get('Scale')),'scale_abbr':clean(r.get('Scale Abbr')),'div':clean(r.get('Div')),'customer':clean(r.get('Customer')).upper() or 'STOCK','master_style':clean(r.get('Master Style')),'style_alias':clean(r.get('Style Alias')),'invoice_descr':clean(r.get('Invoice Descr')),'price':clean(r.get('Price')),'pack_qty':clean(r.get('Pack Qty')),'wh':clean(r.get('Wh')),'style_norm':norm(style),'master_style_norm':norm(r.get('Master Style')),'style_alias_norm':norm(r.get('Style Alias'))}
            w.writerow(row)

    print('Writing SKU Z compact...', flush=True)
    sku_z_headers=['style','clr','sku','size_name','size_num','scale_qty','scale_pack_qty','pack_qty','div','scale','scale_abbr','active','size_norm']
    sku_z_count=0
    with (out_dir/'sku_z.csv').open('w', newline='', encoding='utf-8') as f:
        w=csv.DictWriter(f, fieldnames=sku_z_headers); w.writeheader()
        for r in sku_z_iter:
            sku_z_count += 1
            style=clean(r.get('Style') or r.get('STYLE')).upper()
            clr=clean(r.get('Clr') or r.get('CLR')).upper()
            if not style or not clr or style not in known_styles: continue
            size_name=clean(r.get('Size Name') or r.get('SIZE_NAME'))
            size_num=clean(r.get('Size Num') or r.get('SIZE_NUM'))
            w.writerow({
                'style':style,
                'clr':clr,
                'sku':clean(r.get('Sku') or r.get('SKU')),
                'size_name':size_name,
                'size_num':size_num,
                'scale_qty':clean(r.get('Scale Qty') or r.get('SCALE_QTY')),
                'scale_pack_qty':clean(r.get('Scale Pack Qty') or r.get('SCALE_PACK_QTY')),
                'pack_qty':clean(r.get('Pack Qty') or r.get('PACK_QTY')),
                'div':clean(r.get('Div') or r.get('DIV')),
                'scale':clean(r.get('Scale') or r.get('SCALE')),
                'scale_abbr':clean(r.get('Scale Abbr') or r.get('SCALE_ABBR')),
                'active':clean(r.get('Sku Active') or r.get('SKU_ACTIVE') or r.get('Active') or r.get('ACTIVE')),
                'size_norm':norm(size_name or size_num)
            })

    print('Writing UPC compact...', flush=True)
    upc_headers=['style','clr','clr_desc','clr_abbr','upc','size_name','size_num','sku','div','scale','scale_abbr','price','pack_qty','size_norm']
    upc_count=0
    with (out_dir/'upc.csv').open('w', newline='', encoding='utf-8') as f:
        w=csv.DictWriter(f, fieldnames=upc_headers); w.writeheader()
        for r in upc_iter:
            upc_count += 1
            style=clean(r.get('Style')).upper(); clr=clean(r.get('Clr')).upper(); upc=clean(r.get('Upc No'))
            if not style or not clr or not upc or style not in known_styles: continue
            size_name=clean(r.get('Size Name')); size_num=clean(r.get('Size Num'))
            w.writerow({'style':style,'clr':clr,'clr_desc':clean(r.get('Clr Desc')),'clr_abbr':clean(r.get('Clr Abbr')),'upc':upc,'size_name':size_name,'size_num':size_num,'sku':clean(r.get('Sku')),'div':clean(r.get('Div')),'scale':clean(r.get('Scale')),'scale_abbr':clean(r.get('Scale Abbr')),'price':clean(r.get('Price')),'pack_qty':clean(r.get('Pack Qty')),'size_norm':norm(size_name or size_num)})

    print('Writing colors...', flush=True)
    colors=[]; color_count=0
    for r in color_iter:
        color_count += 1
        code=clean(r.get('Color Code')).upper()
        if code: colors.append({'code':code,'abbr':clean(r.get('Color Abbr')),'description':clean(r.get('Color Description')),'nrf':clean(r.get('Nrf Color No')),'active':clean(r.get('Active'))})
    write_csv(out_dir/'colors.csv', ['code','abbr','description','nrf','active'], colors)

    print('Writing warehouses...', flush=True)
    warehouses=[]; whse_count=0
    for r in whse_iter:
        whse_count += 1
        code=clean(r.get('Wh')).upper()
        if not code: continue
        warehouses.append({'wh':code,'name':clean(r.get('Wh Name')),'type':clean(r.get('Wh Type')),'addr1':clean(r.get('Wh Addr 1')),'addr2':clean(r.get('Wh Addr 2')),'city':clean(r.get('Wh City')),'state':clean(r.get('Wh State')),'postal':clean(r.get('Wh Postal')),'country':clean(r.get('Wh Country')),'active':clean(r.get('Wh Active'))})
    write_csv(out_dir/'warehouses.csv', ['wh','name','type','addr1','addr2','city','state','postal','country','active'], warehouses)

    manifest={'version':9,'source_policy':'official_masters_only','customer_profile_policy':'master_only_all_customers_v1','store_csv_policy':'reject_shifted_columns_preserve_customer_store_keys_v1','size_bucket_policy':'vr_sku_z_size_num_to_qty_szn_v1','customer_cancel_open_lines_policy':'discover_exact_customer_master_column_v1','customer_cancel_open_lines_source_columns':sorted(cancel_open_lines_columns),'counts':{'customers':c_count,'stores':s_count,'malformed_store_rows':malformed_store_rows,'sku_rows':sku_count,'sku_z_rows':sku_z_count,'upc_rows':upc_count,'colors':color_count,'warehouses':whse_count}}
    (out_dir/'manifest.json').write_text(__import__('json').dumps(manifest, indent=2), encoding='utf-8')
    print(f'Wrote compact cache to: {out_dir}', flush=True)
    print('Counts:', manifest['counts'], flush=True)

if __name__=='__main__': main()
