#!/usr/bin/env python3
from __future__ import annotations
import csv,json,re,sys
from pathlib import Path
from openpyxl import load_workbook

def clean(v): return '' if v is None else str(v).strip()
def norm(v): return re.sub(r'[^A-Z0-9]','',clean(v).upper())
EXACT={'CANCELOPENLINES','CANCELOPENLINE','CANCELOPENORDERLINES','CANCELOPENSALESORDERLINES','CXLOPENLINES','CXLOPENLINE'}
def headers(path):
    if path.suffix.lower()=='.csv':
        raw=path.read_bytes(); text=None
        for enc in ('utf-8-sig','cp1252','latin1'):
            try: text=raw.decode(enc); break
            except Exception: pass
        return [clean(x) for x in next(csv.reader(text.splitlines()),[])]
    wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active
    for row in ws.iter_rows(values_only=True):
        vals=[clean(x) for x in row]
        if any(vals):
            return next(csv.reader([vals[0]])) if len(vals)==1 and vals[0].count(',')>3 else vals
    return []
def classify(hs):
    exact=[]; semantic=[]; related=[]
    for h in hs:
        t=norm(h)
        if not t: continue
        if t in EXACT: exact.append(h); continue
        has_cancel=any(x in t for x in ('CANCEL','CNCL','CXL')); has_open='OPEN' in t; has_line=any(x in t for x in ('LINE','LINES','LN','LNS'))
        if has_cancel and has_open and has_line: semantic.append(h); continue
        if any(x in t for x in ('CANCEL','CNCL','CXL','OPEN','BACKORDER','BACKORD','LINE')): related.append(h)
    return {'exact':exact,'semantic':semantic,'related':related}
def main():
    master=Path(sys.argv[1] if len(sys.argv)>1 else 'api/masters').resolve(); out=Path(sys.argv[2] if len(sys.argv)>2 else 'api/training/A2000_V4_7_7_2_CUSTOMER_MASTER_FIELD_INVENTORY.json').resolve()
    candidates=[master/'customer_master.csv.xlsx',master/'customer_master.xlsx',master/'customer_master.csv']; found=[]; errors=[]; cfile=None; cc=None
    for p in candidates:
        if not p.exists(): continue
        try:
            hs=headers(p); cl=classify(hs); found.append({'file':str(p),'header_count':len(hs),'headers':hs,'classification':cl})
            if cfile is None: cfile=p; cc=cl
        except Exception as e: errors.append({'file':str(p),'error':str(e)})
    exact=(cc or {}).get('exact',[]); semantic=(cc or {}).get('semantic',[])
    if len(exact)==1: status='FOUND_EXACT'; selected=exact[0]
    elif len(exact)>1: status='AMBIGUOUS'; selected=None
    elif len(semantic)==1: status='FOUND_SEMANTIC'; selected=semantic[0]
    elif len(semantic)>1: status='AMBIGUOUS'; selected=None
    else: status='NOT_EXPORTED_IN_CURRENT_CUSTOMER_MASTER'; selected=None
    report={'audit':'A2000_V4_7_7_2_CUSTOMER_MASTER_FIELD_INVENTORY','master_dir':str(master),'customer_master_file':str(cfile) if cfile else None,'status':status,'selected_column':selected,'exact_candidates':exact,'semantic_candidates':semantic,'related_customer_master_headers':(cc or {}).get('related',[]),'scanned':found,'errors':errors,'safe_to_repurpose_as_back_order':False,'recommended_action':'Use exact Customer Master column.' if status.startswith('FOUND_') else 'Do not invent or repurpose. Export a Customer Master view that includes Cancel Open Lines.'}
    out.parent.mkdir(parents=True,exist_ok=True); out.write_text(json.dumps(report,indent=2),encoding='utf-8')
    print('COPY THIS RESULT TO CHATGPT'); print('='*116); print('AUDIT=A2000_V4_7_7_2_CUSTOMER_MASTER_FIELD_INVENTORY'); print(f'CUSTOMER_MASTER_FIELD_STATUS={status}'); print('CANCEL_OPEN_LINES_SELECTED_COLUMN='+clean(selected)); print('RELATED_CUSTOMER_MASTER_HEADERS='+'|'.join(report['related_customer_master_headers'])); print('REPURPOSE_AS_BACK_ORDER=false'); print(f'REPORT={out}'); print('A2000_WRITES_PERFORMED=NO'); print('SUPABASE_WRITES_PERFORMED=NO'); print('='*116)
    raise SystemExit(2 if status=='AMBIGUOUS' else 0)
if __name__=='__main__': main()
